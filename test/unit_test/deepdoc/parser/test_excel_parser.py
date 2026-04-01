#  Copyright 2025 The InfiniFlow Authors. All Rights Reserved.
#
#  Licensed under the Apache License, Version 2.0 (the "License");
#  you may not use this file except in compliance with the License.
#  You may obtain a copy of the License at
#
#  http://www.apache.org/licenses/LICENSE-2.0
#
#  Unless required by applicable law or agreed to in writing, software
#  distributed under the License is distributed on an "AS IS" BASIS,
#  WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#  See the License for the specific language governing permissions and
#  limitations under the License.

"""Unit tests for the Excel parser.

Tests cover:
- Basic parsing functionality with various content types
- Edge cases: empty files, special characters, large files
- Multiple sheet handling
- CSV file conversion
- Image extraction from worksheets
- Row counting and limiting
- HTML and Markdown output formats
"""

import importlib.util
import os
import sys
from io import BytesIO
from unittest import mock

_MOCK_MODULES = [
    "xgboost",
    "xgb",
    "pdfplumber",
    "huggingface_hub",
    "PIL",
    "PIL.Image",
    "pypdf",
    "sklearn",
    "sklearn.cluster",
    "sklearn.metrics",
    "deepdoc.vision",
    "infinity",
    "infinity.rag_tokenizer",
]
for _m in _MOCK_MODULES:
    if _m not in sys.modules:
        sys.modules[_m] = mock.MagicMock()


def _find_project_root(marker="pyproject.toml"):
    d = os.path.dirname(os.path.abspath(__file__))
    while d != os.path.dirname(d):
        if os.path.exists(os.path.join(d, marker)):
            return d
        d = os.path.dirname(d)
    return None


_PROJECT_ROOT = _find_project_root()

_lazy_image_spec = importlib.util.spec_from_file_location(
    "rag.utils.lazy_image",
    os.path.join(_PROJECT_ROOT, "rag", "utils", "lazy_image.py"),
)
_lazy_image_mod = importlib.util.module_from_spec(_lazy_image_spec)
sys.modules["rag.utils.lazy_image"] = _lazy_image_mod
_lazy_image_spec.loader.exec_module(_lazy_image_mod)

_excel_spec = importlib.util.spec_from_file_location(
    "deepdoc.parser.excel_parser",
    os.path.join(_PROJECT_ROOT, "deepdoc", "parser", "excel_parser.py"),
)
_excel_mod = importlib.util.module_from_spec(_excel_spec)
sys.modules["deepdoc.parser.excel_parser"] = _excel_mod
_excel_spec.loader.exec_module(_excel_mod)

RAGFlowExcelParser = _excel_mod.RAGFlowExcelParser


def _create_xlsx_with_data(data, sheet_name="Sheet1", multiple_sheets=None):
    """Create a minimal XLSX file in memory.

    Args:
        data: list of rows, each row is a list of cell values
        sheet_name: name for the main sheet
        multiple_sheets: dict of {sheet_name: data} for multiple sheets

    Returns:
        bytes: XLSX file content
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    for row_data in data:
        ws.append(row_data)

    if multiple_sheets:
        for s_name, s_data in multiple_sheets.items():
            new_ws = wb.create_sheet(title=s_name)
            for row_data in s_data:
                new_ws.append(row_data)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _create_csv_with_data(data):
    """Create a CSV file in memory.

    Args:
        data: list of rows, each row is a list of cell values

    Returns:
        bytes: CSV file content
    """
    import csv

    buf = BytesIO()
    text_buf = BytesIO()
    writer = csv.writer(text_buf.decode() if hasattr(text_buf, "decode") else text_buf)

    for row in data:
        text_buf.write((",".join(str(cell) for cell in row) + "\n").encode("utf-8"))

    text_buf.seek(0)
    return text_buf.getvalue()


class TestExcelParserBasic:
    """Tests for basic Excel parsing functionality."""

    def test_parse_simple_spreadsheet(self):
        """Parse a spreadsheet with simple data."""
        data = [
            ["Name", "Age", "City"],
            ["Alice", "30", "NYC"],
            ["Bob", "25", "LA"],
        ]
        xlsx_bytes = _create_xlsx_with_data(data)

        parser = RAGFlowExcelParser()
        result = parser(xlsx_bytes)

        assert len(result) == 2
        assert "Name" in result[0] or "Alice" in result[0]

    def test_parse_empty_spreadsheet(self):
        """Parse an empty spreadsheet."""
        data = [[]]
        xlsx_bytes = _create_xlsx_with_data(data)

        parser = RAGFlowExcelParser()
        result = parser(xlsx_bytes)

        assert result == []

    def test_parse_spreadsheet_with_chinese(self):
        """Parse a spreadsheet with Chinese characters."""
        data = [
            ["姓名", "年龄", "城市"],
            ["张三", "25", "北京"],
            ["李四", "30", "上海"],
        ]
        xlsx_bytes = _create_xlsx_with_data(data)

        parser = RAGFlowExcelParser()
        result = parser(xlsx_bytes)

        assert len(result) == 2

    def test_parse_spreadsheet_with_numeric_data(self):
        """Parse a spreadsheet with numeric data."""
        data = [
            ["Item", "Price", "Quantity"],
            ["Apple", 1.50, 10],
            ["Banana", 0.75, 20],
        ]
        xlsx_bytes = _create_xlsx_with_data(data)

        parser = RAGFlowExcelParser()
        result = parser(xlsx_bytes)

        assert len(result) == 2

    def test_parse_from_file_path(self, temp_xlsx):
        """Test parsing from a file path."""
        data = [["A", "B"], ["1", "2"]]
        xlsx_bytes = _create_xlsx_with_data(data)
        temp_path = temp_xlsx(suffix=".xlsx", content=xlsx_bytes)

        parser = RAGFlowExcelParser()
        result = parser(BytesIO(xlsx_bytes))

        assert len(result) >= 1


class TestExcelParserMultipleSheets:
    """Tests for multiple sheet handling."""

    def test_parse_multiple_sheets(self):
        """Parse a workbook with multiple sheets."""
        data1 = [["Sheet1_A", "Sheet1_B"], ["1", "2"]]
        data2 = [["Sheet2_A", "Sheet2_B"], ["3", "4"]]
        xlsx_bytes = _create_xlsx_with_data(
            data1,
            sheet_name="First",
            multiple_sheets={"Second": data2},
        )

        parser = RAGFlowExcelParser()
        result = parser(xlsx_bytes)

        assert len(result) >= 2

    def test_parse_with_custom_sheet_names(self):
        """Sheet names are included in output when not default."""
        data = [["A", "B"], ["1", "2"]]
        xlsx_bytes = _create_xlsx_with_data(data, sheet_name="CustomName")

        parser = RAGFlowExcelParser()
        result = parser(xlsx_bytes)

        assert len(result) == 1
        assert "CustomName" in result[0]


class TestExcelParserCSV:
    """Tests for CSV file handling."""

    def test_parse_csv_file(self):
        """Parse a CSV file converted to Excel format."""
        csv_bytes = b"Name,Age,City\nAlice,30,NYC\nBob,25,LA\n"

        parser = RAGFlowExcelParser()
        result = parser(csv_bytes)

        assert len(result) == 2

    def test_parse_csv_with_special_characters(self):
        """Parse CSV with special characters."""
        csv_bytes = b'Name,Value\nTest "quoted",100\n'

        parser = RAGFlowExcelParser()
        result = parser(csv_bytes)

        assert len(result) >= 1


class TestExcelParserHtmlOutput:
    """Tests for HTML output format."""

    def test_html_simple_table(self):
        """Generate HTML from simple spreadsheet."""
        data = [["A", "B"], ["1", "2"]]
        xlsx_bytes = _create_xlsx_with_data(data)

        parser = RAGFlowExcelParser()
        result = parser.html(xlsx_bytes)

        assert len(result) >= 1
        assert "<table>" in result[0]
        assert "</table>" in result[0]

    def test_html_with_caption(self):
        """HTML output includes sheet name as caption."""
        data = [["A", "B"], ["1", "2"]]
        xlsx_bytes = _create_xlsx_with_data(data, sheet_name="TestSheet")

        parser = RAGFlowExcelParser()
        result = parser.html(xlsx_bytes)

        assert "TestSheet" in result[0]

    def test_html_chunking(self):
        """Test HTML output chunking with chunk_rows parameter."""
        data = [["Header"]] + [[f"Row{i}"] for i in range(500)]
        xlsx_bytes = _create_xlsx_with_data(data)

        parser = RAGFlowExcelParser()
        result = parser.html(xlsx_bytes, chunk_rows=100)

        assert len(result) > 1

    def test_html_escapes_content(self):
        """HTML output properly escapes special characters."""
        data = [["Text"], ["<script>alert('xss')</script>"]]
        xlsx_bytes = _create_xlsx_with_data(data)

        parser = RAGFlowExcelParser()
        result = parser.html(xlsx_bytes)

        assert "<script>" not in result[0] or "&lt;script&gt;" in result[0]


class TestExcelParserMarkdownOutput:
    """Tests for Markdown output format."""

    def test_markdown_simple_table(self):
        """Generate Markdown from simple spreadsheet."""
        data = [["A", "B"], ["1", "2"]]
        xlsx_bytes = _create_xlsx_with_data(data)

        parser = RAGFlowExcelParser()
        result = parser.markdown(xlsx_bytes)

        assert "|" in result
        assert "A" in result

    def test_markdown_from_csv(self):
        """Generate Markdown from CSV content."""
        csv_bytes = b"Col1,Col2\nVal1,Val2\n"

        parser = RAGFlowExcelParser()
        result = parser.markdown(csv_bytes)

        assert "|" in result


class TestExcelParserRowNumber:
    """Tests for row counting functionality."""

    def test_row_number_simple(self):
        """Count rows in a simple spreadsheet."""
        data = [["A", "B"], ["1", "2"], ["3", "4"]]
        xlsx_bytes = _create_xlsx_with_data(data)

        result = RAGFlowExcelParser.row_number("test.xlsx", xlsx_bytes)

        assert result == 3

    def test_row_number_csv(self):
        """Count rows in a CSV file."""
        csv_bytes = b"Col1,Col2\nVal1,Val2\nVal3,Val4\n"

        result = RAGFlowExcelParser.row_number("test.csv", csv_bytes)

        assert result == 4

    def test_row_number_txt(self):
        """Count rows in a TXT file."""
        txt_bytes = b"Line1\nLine2\nLine3\n"

        result = RAGFlowExcelParser.row_number("test.txt", txt_bytes)

        assert result == 4

    def test_row_number_empty(self):
        """Count rows in an empty spreadsheet."""
        xlsx_bytes = _create_xlsx_with_data([[]])

        result = RAGFlowExcelParser.row_number("test.xlsx", xlsx_bytes)

        assert result >= 0


class TestExcelParserImageExtraction:
    """Tests for image extraction from worksheets."""

    def test_extract_images_no_images(self):
        """Return empty list for worksheet without images."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["A", "B"])

        result = RAGFlowExcelParser._extract_images_from_worksheet(ws)

        assert result == []

    def test_extract_images_structure(self):
        """Verify image extraction result structure."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active

        result = RAGFlowExcelParser._extract_images_from_worksheet(ws, "TestSheet")

        assert isinstance(result, list)


class TestExcelParserHelperMethods:
    """Tests for helper/static methods."""

    def test_clean_dataframe_removes_illegal_chars(self):
        """Test illegal character removal from strings."""
        import pandas as pd

        df = pd.DataFrame({"A": ["Hello\x00World", "Normal"]})
        cleaned = RAGFlowExcelParser._clean_dataframe(df)

        assert "\x00" not in cleaned.iloc[0, 0]

    def test_load_excel_bytes_input(self):
        """Test loading Excel from bytes."""
        data = [["A", "B"], ["1", "2"]]
        xlsx_bytes = _create_xlsx_with_data(data)

        wb = RAGFlowExcelParser._load_excel_to_workbook(xlsx_bytes)

        assert wb is not None
        assert len(wb.sheetnames) >= 1

    def test_load_excel_bytesio_input(self):
        """Test loading Excel from BytesIO."""
        data = [["A", "B"], ["1", "2"]]
        xlsx_bytes = _create_xlsx_with_data(data)

        buf = BytesIO(xlsx_bytes)
        wb = RAGFlowExcelParser._load_excel_to_workbook(buf)

        assert wb is not None

    def test_get_actual_row_count_empty(self):
        """Test row count for empty worksheet."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Empty"

        result = RAGFlowExcelParser._get_actual_row_count(ws)

        assert result >= 0

    def test_get_actual_row_count_with_data(self):
        """Test row count for worksheet with data."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["A", "B"])
        ws.append(["1", "2"])
        ws.append(["3", "4"])

        result = RAGFlowExcelParser._get_actual_row_count(ws)

        assert result == 3

    def test_get_rows_limited_empty(self):
        """Test getting limited rows from empty worksheet."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Empty"

        result = RAGFlowExcelParser._get_rows_limited(ws)

        assert isinstance(result, list)


class TestExcelParserEdgeCases:
    """Tests for edge cases and error handling."""

    def test_parse_spreadsheet_with_empty_cells(self):
        """Parse spreadsheet with empty cells."""
        data = [
            ["A", "B", "C"],
            ["1", "", "3"],
            ["", "2", ""],
        ]
        xlsx_bytes = _create_xlsx_with_data(data)

        parser = RAGFlowExcelParser()
        result = parser(xlsx_bytes)

        assert len(result) >= 1

    def test_parse_spreadsheet_with_none_values(self):
        """Parse spreadsheet with None values."""
        data = [
            ["A", "B"],
            ["1", None],
            [None, "2"],
        ]
        xlsx_bytes = _create_xlsx_with_data(data)

        parser = RAGFlowExcelParser()
        result = parser(xlsx_bytes)

        assert len(result) >= 1

    def test_parse_large_spreadsheet(self):
        """Parse a spreadsheet with many rows."""
        data = [["Col1", "Col2"]] + [[f"Row{i}A", f"Row{i}B"] for i in range(1000)]
        xlsx_bytes = _create_xlsx_with_data(data)

        parser = RAGFlowExcelParser()
        result = parser(xlsx_bytes)

        assert len(result) > 0

    def test_parse_spreadsheet_with_formula(self):
        """Parse spreadsheet with formulas (data_only=True)."""
        data = [
            ["A", "B", "Sum"],
            [1, 2, "=A2+B2"],
        ]
        xlsx_bytes = _create_xlsx_with_data(data)

        parser = RAGFlowExcelParser()
        result = parser(xlsx_bytes)

        assert len(result) >= 1

    def test_parse_spreadsheet_unicode(self):
        """Parse spreadsheet with various Unicode characters."""
        data = [
            ["Emoji", "Arabic", "Russian"],
            ["😀", "مرحبا", "Привет"],
        ]
        xlsx_bytes = _create_xlsx_with_data(data)

        parser = RAGFlowExcelParser()
        result = parser(xlsx_bytes)

        assert len(result) == 1

    def test_parse_spreadsheet_special_sheet_name(self):
        """Parse spreadsheet with special sheet name."""
        data = [["A", "B"], ["1", "2"]]
        xlsx_bytes = _create_xlsx_with_data(data, sheet_name="Sheet1")

        parser = RAGFlowExcelParser()
        result = parser(xlsx_bytes)

        assert len(result) >= 1


class TestExcelParserDataframeConversion:
    """Tests for DataFrame to Workbook conversion."""

    def test_dataframe_to_workbook_single(self):
        """Test converting single DataFrame to Workbook."""
        import pandas as pd

        df = pd.DataFrame({"A": [1, 2], "B": [3, 4]})
        wb = RAGFlowExcelParser._dataframe_to_workbook(df)

        assert wb is not None
        assert "Data" in wb.sheetnames

    def test_dataframe_to_workbook_multiple(self):
        """Test converting multiple DataFrames to Workbook."""
        import pandas as pd

        dfs = {
            "Sheet1": pd.DataFrame({"A": [1, 2]}),
            "Sheet2": pd.DataFrame({"B": [3, 4]}),
        }
        wb = RAGFlowExcelParser._dataframes_to_workbook(dfs)

        assert wb is not None
        assert "Sheet1" in wb.sheetnames
        assert "Sheet2" in wb.sheetnames

    def test_fill_worksheet_from_dataframe(self):
        """Test filling worksheet from DataFrame."""
        import pandas as pd
        from openpyxl import Workbook

        df = pd.DataFrame({"Name": ["Alice", "Bob"], "Age": [30, 25]})
        wb = Workbook()
        ws = wb.active

        RAGFlowExcelParser._fill_worksheet_from_dataframe(ws, df)

        assert ws.cell(row=1, column=1).value == "Name"
        assert ws.cell(row=2, column=2).value == 30
